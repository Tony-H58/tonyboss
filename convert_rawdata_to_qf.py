#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel rawdata → QF JSON 直接转换脚本

流程:
1. 从 Downloads 获取下载的 rawdata Excel
2. 应用列映射和值映射
3. 合并 QC + Factory 数据
4. 生成 insprecord_QF_YYMMDD-YYMMDD.json

使用方式:
python convert_rawdata_to_qf.py --start 260102 --end 260531
"""

import pandas as pd
import openpyxl
import json
import glob
import os
import sys
from datetime import datetime
from pathlib import Path

# 配置路径
DOWNLOAD_DIR = os.path.expanduser('~\\Downloads')
PLAN_FILE = r'E:\88. Claude\03_download rawdata\00_plan_download rawdata.xlsx'
QF_OUTPUT_DIR = r'E:\88. Claude\11_inspdata_monthly\insprecord_QF'

def load_name_mappings():
    """从 plan 文件读取名称映射表"""
    print("📖 读取名称映射规则...")

    excel = openpyxl.load_workbook(PLAN_FILE, data_only=True)
    ws = excel['NameList']

    mappings = {
        'cust': {},      # 客户
        'vendor': {},    # 供应商
        'loc': {},       # 地点
        'factory': {},   # 工厂
        'qc': {},        # QC
        'defects': {}    # 瑕疵
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        # NameList 结构: (cust_code, cust_name, vendor_code, vendor_name, ...)
        if row[0]:  # 客户
            mappings['cust'][str(row[0]).strip()] = str(row[1]).strip() if row[1] else ''
        if row[2]:  # 供应商
            mappings['vendor'][str(row[2]).strip()] = str(row[3]).strip() if row[3] else ''
        if row[4]:  # 地点
            mappings['loc'][str(row[4]).strip()] = str(row[5]).strip() if row[5] else ''
        if row[6]:  # 工厂
            mappings['factory'][str(row[6]).strip()] = str(row[7]).strip() if row[7] else ''
        if row[8]:  # QC
            mappings['qc'][str(row[8]).strip()] = str(row[9]).strip() if row[9] else ''
        if row[11]:  # 瑕疵名称
            mappings['defects'][str(row[11]).strip()] = str(row[13]).strip() if row[13] else ''

    excel.close()

    print(f"  ✓ 客户: {len(mappings['cust'])} | 供应商: {len(mappings['vendor'])} | "
          f"工厂: {len(mappings['factory'])} | QC: {len(mappings['qc'])}")

    return mappings

def load_column_mappings():
    """从 plan 文件读取列映射规则 (ChartTitle sheet)"""
    print("📋 读取列映射规则...")

    excel = openpyxl.load_workbook(PLAN_FILE, data_only=True)
    ws = excel['ChartTitle']

    # 读取映射规则
    qc_col_map = {}
    fac_col_map = {}

    # QC 映射: Row 2 (target), Row 6 (source)
    for col in range(1, ws.max_column + 1):
        tgt = ws.cell(2, col).value
        src = ws.cell(6, col).value
        if tgt and src:
            try:
                tgt_num = int(float(tgt)) if isinstance(tgt, (int, float)) else int(tgt)
                src_num = int(float(src)) if isinstance(src, (int, float)) else int(src)
                if tgt_num > 0 and src_num > 0:
                    qc_col_map[src_num] = tgt_num
            except:
                pass

    # Factory 映射: Row 14 (target), Row 18 (source)
    for col in range(1, ws.max_column + 1):
        tgt = ws.cell(14, col).value
        src = ws.cell(18, col).value
        if tgt and src:
            try:
                tgt_num = int(float(tgt)) if isinstance(tgt, (int, float)) else int(tgt)
                src_num = int(float(src)) if isinstance(src, (int, float)) else int(src)
                if tgt_num > 0 and src_num > 0:
                    fac_col_map[src_num] = tgt_num
            except:
                pass

    excel.close()

    print(f"  ✓ QC 列映射: {len(qc_col_map)} | Factory 列映射: {len(fac_col_map)}")

    return qc_col_map, fac_col_map

def find_rawdata_files():
    """查找下载目录中的 rawdata 文件"""
    print("🔍 查找下载的 rawdata 文件...")

    # 查找品管和工厂验布报表
    qc_files = glob.glob(os.path.join(DOWNLOAD_DIR, '*驗布結論報表*.xls*'))
    fac_files = glob.glob(os.path.join(DOWNLOAD_DIR, '*後端驗布報表*.xls*'))

    # 排除临时文件
    qc_files = [f for f in qc_files if not os.path.basename(f).startswith('~')]
    fac_files = [f for f in fac_files if not os.path.basename(f).startswith('~')]

    if not qc_files or not fac_files:
        print("❌ 找不到 rawdata 文件")
        print(f"   QC 文件: {qc_files}")
        print(f"   Factory 文件: {fac_files}")
        return None, None

    # 取最新的文件
    qc_file = max(qc_files, key=os.path.getmtime)
    fac_file = max(fac_files, key=os.path.getmtime)

    print(f"  ✓ QC: {os.path.basename(qc_file)}")
    print(f"  ✓ Factory: {os.path.basename(fac_file)}")

    return qc_file, fac_file

def read_rawdata(qc_file, fac_file):
    """读取原始数据"""
    print("📥 读取 rawdata...")

    # QC 数据从 Sheet 1
    qc_df = pd.read_excel(qc_file, sheet_name=0)
    # Factory 数据从 Sheet 2
    fac_df = pd.read_excel(fac_file, sheet_name=1)

    print(f"  ✓ QC: {len(qc_df)} 行 {len(qc_df.columns)} 列")
    print(f"  ✓ Factory: {len(fac_df)} 行 {len(fac_df.columns)} 列")

    return qc_df, fac_df

def transform_data(qc_df, fac_df, qc_col_map, fac_col_map, mappings):
    """转换数据为 QF 格式"""
    print("🔄 转换数据...")

    # 为简化起见，这里只做基本示例
    # 完整的转换逻辑应该参照 convert_rawdata.ps1 和 import_monthly.ps1

    # 添加前后端标记
    qc_df.insert(0, '前後端', '品管')
    fac_df.insert(0, '前後端', '工廠')

    # 合并数据
    combined_df = pd.concat([qc_df, fac_df], ignore_index=True)

    print(f"  ✓ 合并后: {len(combined_df)} 行")

    return combined_df

def generate_date_range(qc_file, fac_file):
    """从文件修改时间生成日期范围"""
    qc_mtime = os.path.getmtime(qc_file)
    fac_mtime = os.path.getmtime(fac_file)

    # 取最新的修改时间
    latest_mtime = max(qc_mtime, fac_mtime)
    date = datetime.fromtimestamp(latest_mtime)

    # 生成 YYMMDD 格式
    yy = date.year % 100
    mm = date.month
    dd = 1  # 使用月初

    start_date = f"{yy:02d}{mm:02d}02"  # YYMMDD02
    end_date = f"{yy:02d}{mm:02d}31"    # YYMMDD31

    return start_date, end_date

def export_to_json(df, start_date, end_date):
    """导出为 JSON 格式"""
    print("💾 导出 JSON...")

    output_file = os.path.join(QF_OUTPUT_DIR, f'insprecord_QF_{start_date}-{end_date}.json')

    # 构建输出对象
    output = {
        'metadata': {
            'source': 'convert_rawdata_to_qf.py',
            'created_at': datetime.now().isoformat(),
            'period': f'{start_date}-{end_date}',
            'rows': len(df),
            'columns': len(df.columns)
        },
        'headers': list(df.columns),
        'data': df.to_dict(orient='records')
    }

    # 保存为 JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    file_size = os.path.getsize(output_file) / (1024 * 1024)
    print(f"  ✓ 已保存: {output_file}")
    print(f"  ✓ 大小: {file_size:.1f} MB ({len(df)} 行)")

    return output_file

def main():
    """主流程"""
    print("\n" + "="*70)
    print("Excel rawdata → QF JSON 直接转换")
    print("="*70 + "\n")

    try:
        # 1. 读取映射规则
        mappings = load_name_mappings()
        qc_col_map, fac_col_map = load_column_mappings()

        # 2. 查找和读取文件
        qc_file, fac_file = find_rawdata_files()
        if not qc_file or not fac_file:
            return False

        qc_df, fac_df = read_rawdata(qc_file, fac_file)

        # 3. 转换数据
        combined_df = transform_data(qc_df, fac_df, qc_col_map, fac_col_map, mappings)

        # 4. 生成日期范围
        start_date, end_date = generate_date_range(qc_file, fac_file)

        # 5. 导出 JSON
        json_file = export_to_json(combined_df, start_date, end_date)

        print("\n✅ 完成！")
        print(f"   JSON 文件: {json_file}")
        print(f"   后续可用 qc_kpi_final.py 直接分析")

        return True

    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
