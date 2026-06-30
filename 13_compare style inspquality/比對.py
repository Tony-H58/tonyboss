# -*- coding: utf-8 -*-
"""品管 vs 工廠比對工具"""
import pandas as pd
import os
from datetime import datetime
import webbrowser
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MONTHLY = r"E:\88. Claude\11_inspdata_monthly\insprecord_QF"
OUTPUT = r"E:\88. Claude\13_compare style inspquality"

DESIRED_COLUMNS = [
    '前後端', '實際驗布日', '客戶', '款號', '採購人員', '供應商', '供應地',
    '採購單號', '成衣產區', 'BL NO', '成份', '針平織', '布種', '染法', 'QC',
    '顏色', '收料碼數', '抽驗碼數', 'A級碼數', 'B級碼數', 'C級碼數', '總瑕疵點數',
    '布面瑕疵', '外觀問題', '手感問題', '緯斜問題', '顏色問題', '規格問題',
    '其他問題', 'C20需出貨原因', 'C20解決方案', '簽核狀態'
]

def parse_month(month_str):
    """解析月份字符串，返回日期范围"""
    month_str = month_str.strip()
    if '月' in month_str:
        month = int(month_str.replace('月', '').strip())
        return f"2026-{month:02d}-01", f"2026-{month:02d}-28"
    return None, None

def load_data(query_value, date_from, date_to):
    """加载指定条件和日期范围的数据（支持款号或采购单号）"""
    dfs = []

    for p in [MONTHLY]:
        if os.path.exists(p):
            for f in os.listdir(p):
                if f.endswith(('.csv', '.xlsx')):
                    try:
                        if f.endswith('.csv'):
                            df = pd.read_csv(os.path.join(p, f), encoding='utf-8-sig')
                        else:
                            df = pd.read_excel(os.path.join(p, f))
                        dfs.append(df)
                    except:
                        pass

    if not dfs:
        return None

    df = pd.concat(dfs, ignore_index=True).drop_duplicates()

    # 筛选款号或采购单号
    if query_value.startswith('TMKF'):
        # 采购单号
        po_cols = [c for c in df.columns if '採購單' in str(c)]
        if po_cols:
            df = df[df[po_cols[0]].astype(str).str.contains(query_value, case=False, na=False)]
    else:
        # 款号
        style_cols = [c for c in df.columns if '款' in str(c)]
        if style_cols:
            df = df[df[style_cols[0]].astype(str).str.contains(query_value, case=False, na=False)]

    # 筛选日期
    date_cols = [c for c in df.columns if '日' in str(c)]
    if date_cols:
        df[date_cols[0]] = pd.to_datetime(df[date_cols[0]], errors='coerce')
        df = df[(df[date_cols[0]] >= date_from) & (df[date_cols[0]] <= date_to)]

    # 只保留指定欄位
    keep_cols = [c for c in DESIRED_COLUMNS if c in df.columns]
    df = df[keep_cols]

    return df

def compare(df):
    """比对品管和工厂数据"""
    if df is None or len(df) == 0:
        return None

    # 分开品管（前端）和工厂（后端）
    front_back_cols = [c for c in df.columns if '前後' in str(c) or '前后' in str(c)]
    if not front_back_cols:
        return df

    col = front_back_cols[0]
    qc = df[df[col].astype(str).str.contains('前|品管', case=False, na=False, regex=True)]
    fty = df[df[col].astype(str).str.contains('後|后|工廠|工厂', case=False, na=False, regex=True)]

    return {
        'all': df,
        'qc': qc,
        'fty': fty,
        'summary': pd.DataFrame({
            '類別': ['品管(前端)', '工廠(後端)'],
            '筆數': [len(qc), len(fty)]
        }),
        'by_color': compare_by_color(qc, fty)
    }

def compare_by_color(qc, fty):
    """按顏色彙總，品管/工廠並排"""
    def agg(d, prefix):
        if len(d) == 0 or '顏色' not in d.columns:
            return pd.DataFrame(columns=['顏色', f'{prefix}碼數', f'{prefix}筆數', f'{prefix}C級%'])
        d = d.copy()
        d['C級%'] = (d['C級碼數'] / d['抽驗碼數'].replace(0, pd.NA) * 100).fillna(0)
        g = d.groupby('顏色').agg(**{
            f'{prefix}碼數': ('收料碼數', 'sum'),
            f'{prefix}筆數': ('顏色', 'count'),
            f'{prefix}C級%': ('C級%', 'mean')
        }).reset_index()
        g[f'{prefix}C級%'] = g[f'{prefix}C級%'].round(2)
        return g

    qc_g = agg(qc, '品管')
    fty_g = agg(fty, '工廠')
    return pd.merge(qc_g, fty_g, on='顏色', how='outer').fillna(0)

# 主程式
import sys

print("="*50)
print("品管 vs 工廠比對工具")
print("="*50 + "\n")

# 输入条件 - 支持命令行参数或交互式输入
if len(sys.argv) > 2:
    query = sys.argv[1]
    month_input = sys.argv[2]
else:
    query = input("款號或採購單號 (如 MS6FK213R_FA26 或 TMKF-26-04485): ").strip()
    month_input = input("月份 (如 5月): ").strip()

date_from, date_to = parse_month(month_input)
if not date_from:
    print("日期格式错误")
    exit(1)

print(f"\n查詢條件: {query}, 日期範圍={date_from}~{date_to}\n")

# 加载和比对
print("載入資料...")
df = load_data(query, date_from, date_to)

if df is None or len(df) == 0:
    print("查無資料")
    exit(1)

print(f"加載 {len(df)} 筆記錄\n")

result = compare(df)

if result is None:
    print("比對失敗")
    exit(1)

# 显示摘要
print(result['summary'].to_string(index=False))
print()

# 保存到 Excel（单页，清除旧数据保留标题）
output_path = os.path.join(OUTPUT, "00_analysis data.xlsx")

# 如果文件存在，先清除内容（只保留标题）
if os.path.exists(output_path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        if '驗布記錄' in wb.sheetnames:
            ws = wb['驗布記錄']
        else:
            ws = wb.active
        # 保留标题，删除数据行
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        wb.save(output_path)
    except:
        pass

# 写入新数据
result['all'].to_excel(output_path, index=False, sheet_name='驗布記錄', startrow=0)
print(f"已保存: {output_path}\n")

# 生成 HTML（空值顯示為空白，不顯示 NaN）
by_color = result['by_color']
if len(by_color) > 0:
    total_row = {'顏色': '總計'}
    for c in by_color.columns:
        if c == '顏色':
            continue
        if c.endswith('%'):
            total_row[c] = round(by_color[c].mean(), 2)
        else:
            total_row[c] = by_color[c].sum()
    by_color = pd.concat([by_color, pd.DataFrame([total_row])], ignore_index=True)

def format_date_col(df):
    if '實際驗布日' in df.columns:
        df = df.copy()
        df['實際驗布日'] = pd.to_datetime(df['實際驗布日'], errors='coerce').dt.strftime('%Y-%m-%d')
    return df

by_color_html = by_color.fillna('') if len(by_color) > 0 else None
qc_html = format_date_col(result['qc'].head(20)).fillna('') if len(result['qc']) > 0 else None
fty_html = format_date_col(result['fty'].head(20)).fillna('') if len(result['fty']) > 0 else None

import html as html_lib

COLUMN_WIDTHS = {
    '實際驗布日': '7em',
    '採購單號': '15ch',
    '款號': '18ch',
    '成衣產區': '11ch',
}

def df_to_html_table(df, extra_widths=None):
    """產生固定欄寬/高、文字過長省略並顯示title提示的表格"""
    widths = {**COLUMN_WIDTHS, **(extra_widths or {})}

    def width_style(c):
        w = widths.get(c)
        return f' style="width:{w}"' if w else ''

    ths = ''.join(f'<th{width_style(c)}>{html_lib.escape(str(c))}</th>' for c in df.columns)
    rows = ''
    for _, row in df.iterrows():
        tds = ''
        for c, v in zip(df.columns, row):
            w = widths.get(c)
            td_style = f' style="width:{w}"' if w else ''
            cell_style = f' style="width:{w}"' if w else ''
            tds += f'<td{td_style} title="{html_lib.escape(str(v))}"><div class="cell"{cell_style}>{html_lib.escape(str(v))}</div></td>'
        rows += f'<tr>{tds}</tr>'
    return f'<table><thead><tr>{ths}</tr></thead><tbody>{rows}</tbody></table>'

html = f"""<html><head><meta charset='utf-8'><style>
body{{font-family:Arial;padding:20px;background:#f5f5f5}}
.container{{max-width:1000px;margin:0 auto;background:white;padding:20px;border-radius:8px}}
h1{{color:#333;font-size:1.6em}}
table{{border-collapse:collapse;width:100%;margin:20px 0;table-layout:fixed}}
th{{background:#667eea;color:white;padding:8px;text-align:left;height:32px;overflow:hidden;width:5em}}
td{{padding:0;border-bottom:1px solid #ddd;height:32px;width:5em}}
td .cell{{padding:8px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;height:32px;line-height:16px;box-sizing:border-box;width:5em}}
.summary{{background:#f0f0f0;padding:15px;border-radius:6px;margin:20px 0}}
.detail-small table{{font-size:8px}}
.detail-small th{{padding:4px;height:16px}}
.detail-small td{{height:16px}}
.detail-small td .cell{{padding:4px;height:16px;line-height:8px}}
.detail-small h3{{font-size:80%}}
.color-stats{{font-size:80%}}
</style></head><body>
<div class='container'>
<h1>比對結果: {query}</h1>
<div class='color-stats'>
<h3>按顏色統計（品管/工廠並排）</h3>
{df_to_html_table(by_color_html, extra_widths={'顏色': '10em'}) if by_color_html is not None else '無資料'}
</div>
<div class='detail-small'>
<h3>品管 ({len(result['qc'])} 筆)</h3>
{df_to_html_table(qc_html) if qc_html is not None else '無資料'}
<h3>工廠 ({len(result['fty'])} 筆)</h3>
{df_to_html_table(fty_html) if fty_html is not None else '無資料'}
</div>
</div></body></html>"""

html_path = os.path.join(OUTPUT, "compare_result.html")
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"已生成: {html_path}\n")

# 自動複製到 00_html file/13_compare style inspquality/
try:
    import shutil
    backup_dir = r"E:\88. Claude\00_html file\13_compare style inspquality"
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, "compare_result.html")
    shutil.copy2(html_path, backup_path)
    print(f"✓ 已複製到: {backup_path}\n")
except Exception as e:
    print(f"⚠ 複製失敗: {e}\n")

webbrowser.open(f'file:///{html_path}')

print("完成!")
